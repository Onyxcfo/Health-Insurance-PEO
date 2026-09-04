# -*- coding: utf-8 -*-
"""Onyx Accounting Group - 2026 medical plan comparison workbook.
Vendors: ADP TotalSource, TriNet, Angle Health/CapFi.  Effective 10/1/2026.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from plans import PLANS, FLOORS, TIERS

F   = "Arial"
INK = Font(name=F, size=10)
BLD = Font(name=F, size=10, bold=True)
BLU = Font(name=F, size=10, color="0000FF")           # hardcoded input
BLUB= Font(name=F, size=10, bold=True, color="0000FF")
GRN = Font(name=F, size=10, color="008000")           # cross-sheet link
TTL = Font(name=F, size=14, bold=True)
SUB = Font(name=F, size=9, italic=True, color="595959")
SEC = Font(name=F, size=11, bold=True, color="1D5B72")
HDF = Font(name=F, size=9, bold=True, color="FFFFFF")
HFILL = PatternFill("solid", fgColor="1D5B72")
YEL   = PatternFill("solid", fgColor="FFFF00")
BAND  = PatternFill("solid", fgColor="EFF3F5")
thin  = Side(style="thin", color="BFBFBF")
BOX   = Border(left=thin, right=thin, top=thin, bottom=thin)
BOT   = Border(bottom=Side(style="thin", color="1D5B72"))

CUR  = '$#,##0;($#,##0);-'
CUR2 = '$#,##0.00;($#,##0.00);-'
PCT  = '0.0%'

wb = openpyxl.Workbook()

# =====================================================================
# 1. ASSUMPTIONS
# =====================================================================
a = wb.active
a.title = "Assumptions"

def sec(ws, row, text):
    c = ws.cell(row=row, column=1, value=text); c.font = SEC; c.border = BOT
    for col in range(2, 8):
        ws.cell(row=row, column=col).border = BOT

def hdr(ws, row, labels, start=1):
    for i, l in enumerate(labels):
        c = ws.cell(row=row, column=start+i, value=l)
        c.font = HDF; c.fill = HFILL; c.border = BOX
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

a["A1"] = "Onyx Accounting Group, LLC - 2026 Medical Plan Comparison"
a["A1"].font = TTL
a["A2"] = "All plans offered by ADP TotalSource, TriNet and Angle Health / CapFi.  Medical effective 10/1/2026.  Dental, vision, life and disability are priced separately and are NOT in this workbook."
a["A2"].font = SUB

a["A4"] = "LEGEND"; a["A4"].font = BLD
for r, (col, txt) in enumerate([
    (BLU, "Blue text = hardcoded input taken from a vendor document or from Josephine.  These are the only cells to edit."),
    (INK, "Black text = formula.  Recalculates from the blue inputs."),
    (GRN, "Green text = link to another tab in this workbook."),
], start=5):
    a.cell(row=r, column=1, value="").fill = YEL if r == 5 else PatternFill()
    c = a.cell(row=r, column=2, value=txt); c.font = col
a["A5"].fill = YEL

sec(a, 8, "1.  ONYX EMPLOYER CONTRIBUTION TARGET  (fixed dollars, medical only)")
hdr(a, 9, ["Tier", "Code", "Onyx target $/mo", "Onyx target $/yr", "Notes"])
for i, (name, code, amt) in enumerate(TIERS):
    r = 10 + i
    a.cell(row=r, column=1, value=name).font = INK
    a.cell(row=r, column=2, value=code).font = BLU
    c = a.cell(row=r, column=3, value=amt); c.font = BLU; c.number_format = CUR2; c.fill = YEL
    c = a.cell(row=r, column=4, value=f"=C{r}*12"); c.font = INK; c.number_format = CUR
    note = {"EE": "Set by Onyx.  Applies to Lisa.",
            "ES": "Set by Onyx.  Applies to Christine.",
            "EC": "ASSUMPTION - no target was stated for this tier and nobody occupies it today.  Defaulted to the employee+spouse figure.  Change if wrong.",
            "FAM": "Set by Onyx.  Applies to Steven and Jessica."}[code]
    a.cell(row=r, column=5, value=note).font = SUB
    for col in range(1, 6):
        a.cell(row=r, column=col).border = BOX
a["A14"] = "Target cost at current enrollment (4 enrolled)"; a["A14"].font = BLD
a["C14"] = "=SUMIFS(C10:C13,B10:B13,\"EE\")*1+SUMIFS(C10:C13,B10:B13,\"ES\")*1+SUMIFS(C10:C13,B10:B13,\"FAM\")*2"
a["C14"].font = BLD; a["C14"].number_format = CUR2
a["D14"] = "=C14*12"; a["D14"].font = BLD; a["D14"].number_format = CUR

sec(a, 16, "2.  VENDOR EMPLOYER CONTRIBUTION FLOORS  (minimum Onyx must pay per enrolled employee, applied flat to every tier)")
hdr(a, 17, ["Vendor", "% required", "Reference plan employee-only premium", "Floor $/mo per enrolled EE", "Reference plan", "Source / status"])
fl_src = {
 "ADP":    "Stated in writing: 50% of the lowest-cost plan, employee-only rate.",
 "TriNet": "CONFIRMED by TriNet: 70% of the minimum funding on the lowest-cost plan offered.  Our derivation was right and reconciles to the penny on both carriers quoted.  Keyed to the ACO 6500's $385 employee-only rate.",
 "Angle":  "Level-funded minimum participation funding; 50% of the lowest employee-only rate.",
}
for i, (v, pct, ref, refplan) in enumerate(FLOORS):
    r = 18 + i
    a.cell(row=r, column=1, value=v).font = INK
    c = a.cell(row=r, column=2, value=pct); c.font = BLU; c.number_format = PCT; c.fill = YEL
    c = a.cell(row=r, column=3, value=ref); c.font = BLU; c.number_format = CUR2; c.fill = YEL
    c = a.cell(row=r, column=4, value=f"=B{r}*C{r}"); c.font = INK; c.number_format = CUR2
    a.cell(row=r, column=5, value=refplan).font = INK
    a.cell(row=r, column=6, value=fl_src[v]).font = SUB
    for col in range(1, 7):
        a.cell(row=r, column=col).border = BOX
a["A22"] = "The floor only bites where it exceeds the tier target.  At Onyx's targets it never does - every floor here is below $300 - so Onyx's stated dollars govern on all three vendors."
a["A22"].font = SUB
a["A23"] = "Insperity is NOT in this workbook.  Its floor mechanic is different (75% of a designated Base Plan's employee-only rate: $509.76 Saver / $646.90 Select) and it exceeds every tier target, so it cannot be compared on the same grid."
a["A23"].font = SUB

sec(a, 25, "3.  ENROLLMENT USED FOR THE TOTAL-COST TAB")
hdr(a, 26, ["Employee", "Annual pay", "Medical tier", "Code", "Enrolling?", "Note"])
enr = [("Steven Nikolov", 186000, "Family", "FAM", "Yes", "100% owner.  >2% S-corp shareholder - premiums cannot run pretax through the cafeteria plan (IRC 1372).  Whether he enrolls is still open; worth $12,000/yr of Onyx cost."),
       ("Lisa Danforth", 178094, "Employee only", "EE", "Yes", ""),
       ("Jessica B.", 135000, "Family", "FAM", "Yes", "New hire."),
       ("Christine Johnson", 79228, "Employee + spouse", "ES", "Yes", ""),
       ("Josephine Mack", 80000, "Family", "FAM", "No", "Waiving medical, dental and vision.  Shown for reference only - if she enrolled it would be family coverage, +$12,000/yr to Onyx at every vendor.")]
for i, (n, pay, tier, code, y, note) in enumerate(enr):
    r = 27 + i
    a.cell(row=r, column=1, value=n).font = INK
    c = a.cell(row=r, column=2, value=pay); c.font = BLU; c.number_format = CUR
    a.cell(row=r, column=3, value=tier).font = INK
    a.cell(row=r, column=4, value=code).font = BLU
    c = a.cell(row=r, column=5, value=y); c.font = BLU; c.fill = YEL
    c.alignment = Alignment(horizontal="center")
    a.cell(row=r, column=6, value=note).font = SUB
    for col in range(1, 7):
        a.cell(row=r, column=col).border = BOX
a["A32"] = "Elena Nikolov is not on payroll and is not an enrollee."; a["A32"].font = SUB

sec(a, 34, "4.  HOW EACH COLUMN IS CALCULATED")
meth = [
 ("Onyx pays $/mo", "Formula: MIN(premium, MAX(tier target, vendor floor)).  The MIN stops Onyx from being shown as paying more than the premium; it does not bind on any plan here."),
 ("Employee pays $/mo", "Formula: premium minus Onyx pays.  This is the payroll deduction."),
 ("Worst case, employee", "Formula: employee annual premium plus the household out-of-pocket maximum for that tier.  Premiums never count toward an out-of-pocket maximum, so the two add without overlapping.  Employee-only rows use the individual OOP max; all other tiers use the family OOP max."),
 ("Out-of-pocket maximum", "In-network only.  Every plan here covers preventive care at 100% before the deductible, so the worst case assumes a full year of non-preventive claims."),
 ("Coinsurance", "Shown as the share the PLAN pays after the deductible.  100% means the plan pays everything after the deductible is met.  Aetna's own documents state the member's share; those were converted."),
 ("'ded' / 'd' on a copay", "That copay or coinsurance applies only AFTER the deductible is satisfied."),
 ("HSA-eligible", "Y only where the plan is designated an HDHP and charges no non-preventive copay before the deductible.  HSA contributions are deliberately NOT modeled - Onyx cannot know what any employee will choose to fund, so a plan's true cost to an employee with an HSA is lower than the worst case shown."),
 ("Premiums", "Composite rates as quoted for Onyx's actual census.  Quote IDs: TriNet Q-00412887, Angle 400798.  ADP rates from the TotalSource 2026 renewal deck."),
 ("Not modeled", "Admin fees, workers' compensation, setup fees, dental, vision, life and disability.  Those sit in the employer-only cost table in the Steven report, not here.  This workbook is medical premium only."),
]
hdr(a, 35, ["Item", "Treatment"])
for i, (k, v) in enumerate(meth):
    r = 36 + i
    a.cell(row=r, column=1, value=k).font = BLD
    a.cell(row=r, column=2, value=v).font = INK
    a.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    for col in (1, 2):
        a.cell(row=r, column=col).border = BOX
    a.row_dimensions[r].height = 30

sec(a, 46, "5.  OPEN ITEMS THAT COULD MOVE THESE NUMBERS")
for i, t in enumerate([
  "TriNet says TWELVE plans are available; the strategy appendix we were given lists the SIXTEEN Aetna rungs in this workbook.  Which four are unavailable is unknown, so all sixteen are kept here.  It matters only if the ACO 6500 is not among the twelve - the $269.50 floor is keyed to its $385 rate, and dropping the four cheapest would push the floor to $364.00, above the $300 employee-only target.",
  "TriNet has confirmed its service fee is the $1,200/month minimum ($14,400/yr), its workers' compensation is $300, basic life at $10,000 is $67, and nothing else is required - no dental, no vision, and long-term disability is optional.",
  "TriNet allows different funding tiers by class, so Onyx could fund Steven at 100% while holding others at the fixed dollars.  Not modeled here - this workbook applies the stated tier targets uniformly.",
  "Whether Steven enrolls.  Family coverage for him is roughly $12,000/yr of Onyx contribution.",
  "No contribution target has been set for employee + child(ren).  Defaulted to $600.",
  "Angle's renewal exposure.  Level-funded; a 40% increase at renewal is considered likely.",
  "Renewal anchoring differs by vendor: TriNet holds 10/1 in perpetuity, ADP re-rates in June, Insperity renews on the anniversary.  Angle is level-funded and re-underwrites.",
]):
    a.cell(row=47+i, column=1, value=u"•  " + t).font = INK

for col, w in [("A", 34), ("B", 46), ("C", 20), ("D", 18), ("E", 34), ("F", 60)]:
    a.column_dimensions[col].width = w
a.sheet_view.showGridLines = False
a.freeze_panes = "A3"

# =====================================================================
# 2. PLAN DESIGN
# =====================================================================
p = wb.create_sheet("Plan Design")
p["A1"] = "Plan design - all 30 plans"
p["A1"].font = TTL
p["A2"] = "Monthly premiums are composite rates quoted for Onyx's census.  Coinsurance is the share the PLAN pays after the deductible.  'ded' or 'd' means the copay applies only after the deductible."
p["A2"].font = SUB
PD_HDR = ["Vendor", "Plan", "Key", "Network", "HSA", "Ded individual", "Ded family",
          "OOP max individual", "OOP max family", "Coins (plan pays)", "PCP visit",
          "Specialist", "Rx generic / preferred / non-preferred", "Emergency room",
          "EE", "ES", "EC", "FAM"]
p["O3"] = "Monthly premium by tier"; p["O3"].font = BLD
p.merge_cells("O3:R3")
p["O3"].alignment = Alignment(horizontal="center")
hdr(p, 4, PD_HDR)
p.row_dimensions[4].height = 42
PD_FIRST = 5
for i, row in enumerate(PLANS):
    r = PD_FIRST + i
    v, plan, net, hsa, di, df, oi, of, cp, pcp, spec, rx, er, ee, es, ec, fam = row
    vals = [v, plan, f"{v}|{plan}", net, hsa, di, df, oi, of, cp, pcp, spec, rx, er, ee, es, ec, fam]
    for j, val in enumerate(vals):
        c = p.cell(row=r, column=1+j, value=val)
        c.font = BLU
        c.border = BOX
        if j in (5, 6, 7, 8):
            c.number_format = CUR
        if j in (14, 15, 16, 17):
            c.number_format = CUR2
        if j == 9:
            c.number_format = "0%"
        if j in (4, 9):
            c.alignment = Alignment(horizontal="center")
    if i % 2:
        for j in range(18):
            p.cell(row=r, column=1+j).fill = BAND
PD_LAST = PD_FIRST + len(PLANS) - 1
p.cell(row=PD_LAST+2, column=1, value="Every cell above is a hardcoded input transcribed from the vendor's own document.  Sources: ADP TotalSource 2026 renewal deck pp.6-8; TriNet quote Q-00412887 strategy appendix pp.13-15; Angle Health / CapFi presentation quote 400798.").font = SUB
p.cell(row=PD_LAST+3, column=1, value="Angle out-of-network: 50% coinsurance after a separate out-of-network deductible.  ADP EPO plans have no out-of-network benefit at all except emergencies.").font = SUB
for col, w in [("A", 9), ("B", 30), ("C", 36), ("D", 34), ("E", 6), ("F", 13), ("G", 13),
               ("H", 14), ("I", 14), ("J", 12), ("K", 15), ("L", 15), ("M", 26), ("N", 15),
               ("O", 11), ("P", 11), ("Q", 11), ("R", 11)]:
    p.column_dimensions[col].width = w
p.column_dimensions["C"].hidden = True
p.freeze_panes = "C5"
p.sheet_view.showGridLines = False

# =====================================================================
# 3. COST BY TIER
# =====================================================================
c3 = wb.create_sheet("Cost by Tier")
c3["A1"] = "What each plan costs Onyx and the employee, by tier"
c3["A1"].font = TTL
c3["A2"] = "One row per plan per tier.  Onyx pays the greater of its tier target and the vendor's floor, never more than the premium.  Worst case = the employee's annual premium plus the household out-of-pocket maximum."
c3["A2"].font = SUB
CT_HDR = ["Vendor", "Plan", "Key", "Tier", "Code", "HSA", "Monthly premium",
          "Onyx tier target $/mo", "Vendor floor $/mo", "Onyx pays $/mo",
          "Employee pays $/mo", "Onyx pays $/yr", "Employee pays $/yr",
          "Onyx share of premium", "Deductible (this tier)", "OOP max (this tier)",
          "Worst case to employee $/yr", "Row key"]
hdr(c3, 4, CT_HDR)
c3.row_dimensions[4].height = 42

PDK  = f"'Plan Design'!$C${PD_FIRST}:$C${PD_LAST}"
PDPR = f"'Plan Design'!$O${PD_FIRST}:$R${PD_LAST}"
PDTH = f"'Plan Design'!$O$4:$R$4"
PDDI = f"'Plan Design'!$F${PD_FIRST}:$F${PD_LAST}"
PDDF = f"'Plan Design'!$G${PD_FIRST}:$G${PD_LAST}"
PDOI = f"'Plan Design'!$H${PD_FIRST}:$H${PD_LAST}"
PDOF = f"'Plan Design'!$I${PD_FIRST}:$I${PD_LAST}"
PDHS = f"'Plan Design'!$E${PD_FIRST}:$E${PD_LAST}"

CT_FIRST = 5
r = CT_FIRST
for row in PLANS:
    v, plan = row[0], row[1]
    key = f"{v}|{plan}"
    for tname, tcode, _t in TIERS:
        c3.cell(row=r, column=1, value=v).font = INK
        c3.cell(row=r, column=2, value=plan).font = INK
        c3.cell(row=r, column=3, value=key).font = INK
        c3.cell(row=r, column=4, value=tname).font = INK
        c3.cell(row=r, column=5, value=tcode).font = INK
        f = {}
        f["F"] = f"=INDEX({PDHS},MATCH($C{r},{PDK},0))"
        f["G"] = f"=INDEX({PDPR},MATCH($C{r},{PDK},0),MATCH($E{r},{PDTH},0))"
        f["H"] = f"=INDEX(Assumptions!$C$10:$C$13,MATCH($E{r},Assumptions!$B$10:$B$13,0))"
        f["I"] = f"=INDEX(Assumptions!$D$18:$D$20,MATCH($A{r},Assumptions!$A$18:$A$20,0))"
        f["J"] = f"=MIN($G{r},MAX($H{r},$I{r}))"
        f["K"] = f"=$G{r}-$J{r}"
        f["L"] = f"=$J{r}*12"
        f["M"] = f"=$K{r}*12"
        f["N"] = f"=IF($G{r}=0,0,$J{r}/$G{r})"
        f["O"] = f'=IF($E{r}="EE",INDEX({PDDI},MATCH($C{r},{PDK},0)),INDEX({PDDF},MATCH($C{r},{PDK},0)))'
        f["P"] = f'=IF($E{r}="EE",INDEX({PDOI},MATCH($C{r},{PDK},0)),INDEX({PDOF},MATCH($C{r},{PDK},0)))'
        f["Q"] = f"=$M{r}+$P{r}"
        f["R"] = f"=$C{r}&\"|\"&$E{r}"
        for col, formula in f.items():
            cell = c3[f"{col}{r}"]
            cell.value = formula
            cell.font = GRN if col in ("F", "G", "H", "I", "O", "P") else INK
        for col in ("G", "H", "I", "J", "K"):
            c3[f"{col}{r}"].number_format = CUR2
        for col in ("L", "M", "O", "P", "Q"):
            c3[f"{col}{r}"].number_format = CUR
        c3[f"N{r}"].number_format = PCT
        c3[f"F{r}"].alignment = Alignment(horizontal="center")
        c3[f"Q{r}"].font = Font(name=F, size=10, bold=True)
        c3[f"R{r}"].font = INK
        for col in range(1, 18):
            c3.cell(row=r, column=col).border = BOX
        if tcode in ("ES", "FAM"):
            for col in range(1, 18):
                c3.cell(row=r, column=col).fill = BAND
        r += 1
CT_LAST = r - 1
c3.cell(row=CT_LAST+2, column=1, value="Worst case is the true ceiling for that household in one plan year: it assumes the family hits the in-network out-of-pocket maximum in full.  It does not include out-of-network care, balance billing, or anything the plan excludes.").font = SUB
c3.cell(row=CT_LAST+3, column=1, value="HSA funding is not modeled.  On an HSA-eligible plan an employee can pay part of that worst case with pretax dollars, so the real cost is lower - but by an amount only the employee controls.").font = SUB
for col, w in [("A", 9), ("B", 30), ("C", 36), ("D", 21), ("E", 6), ("F", 6), ("G", 12),
               ("H", 13), ("I", 12), ("J", 12), ("K", 13), ("L", 12), ("M", 13),
               ("N", 12), ("O", 14), ("P", 14), ("Q", 15), ("R", 36)]:
    c3.column_dimensions[col].width = w
c3.column_dimensions["C"].hidden = True
c3.column_dimensions["R"].hidden = True
c3.freeze_panes = "F5"
c3.sheet_view.showGridLines = False
c3.auto_filter.ref = f"A4:Q{CT_LAST}"

# =====================================================================
# 4. ONYX TOTAL COST  (30 plans at actual enrollment)
# =====================================================================
t = wb.create_sheet("Onyx Total Cost")
t["A1"] = "Total cost at Onyx's actual enrollment, one plan at a time"
t["A1"].font = TTL
t["A2"] = "Assumes all four enrollees take the same plan: Steven family, Jessica family, Lisa employee only, Christine employee + spouse.  Combined worst case assumes every household hits its out-of-pocket maximum in the same year - a ceiling, not a forecast."
t["A2"].font = SUB
T_HDR = ["Vendor", "Plan", "Key", "HSA", "Onyx $/mo", "Onyx $/yr",
         "Employees $/mo (all 4)", "Employees $/yr (all 4)",
         "Onyx + employees $/yr", "Combined worst case $/yr",
         "Onyx variance vs $34,800 target"]
hdr(t, 4, T_HDR)
t.row_dimensions[4].height = 42
CK = f"'Cost by Tier'!$C${CT_FIRST}:$C${CT_LAST}"
CE = f"'Cost by Tier'!$E${CT_FIRST}:$E${CT_LAST}"
def s(colletter, tier):
    return f"SUMIFS('Cost by Tier'!${colletter}${CT_FIRST}:${colletter}${CT_LAST},{CK},$C{{r}},{CE},\"{tier}\")"
T_FIRST = 5
for i, row in enumerate(PLANS):
    r = T_FIRST + i
    v, plan = row[0], row[1]
    t.cell(row=r, column=1, value=v).font = INK
    t.cell(row=r, column=2, value=plan).font = INK
    t.cell(row=r, column=3, value=f"{v}|{plan}").font = INK
    t[f"D{r}"] = f"=INDEX({PDHS},MATCH($C{r},{PDK},0))"
    t[f"D{r}"].font = GRN; t[f"D{r}"].alignment = Alignment(horizontal="center")
    t[f"E{r}"] = "=" + s("J", "EE").format(r=r) + "+" + s("J", "ES").format(r=r) + "+2*" + s("J", "FAM").format(r=r)
    t[f"F{r}"] = f"=E{r}*12"
    t[f"G{r}"] = "=" + s("K", "EE").format(r=r) + "+" + s("K", "ES").format(r=r) + "+2*" + s("K", "FAM").format(r=r)
    t[f"H{r}"] = f"=G{r}*12"
    t[f"I{r}"] = f"=F{r}+H{r}"
    t[f"J{r}"] = "=" + s("Q", "EE").format(r=r) + "+" + s("Q", "ES").format(r=r) + "+2*" + s("Q", "FAM").format(r=r)
    t[f"K{r}"] = f"=F{r}-Assumptions!$D$14"
    for col in ("E", "G"):
        t[f"{col}{r}"].number_format = CUR2; t[f"{col}{r}"].font = INK
    for col in ("F", "H", "I", "J", "K"):
        t[f"{col}{r}"].number_format = CUR; t[f"{col}{r}"].font = INK
    t[f"F{r}"].font = BLD
    for col in range(1, 12):
        t.cell(row=r, column=col).border = BOX
    if i % 2:
        for col in range(1, 12):
            t.cell(row=r, column=col).fill = BAND
T_LAST = T_FIRST + len(PLANS) - 1
t.cell(row=T_LAST+2, column=1, value="Onyx's cost is flat across every plan within a vendor, because the contribution is a fixed dollar amount and no vendor's floor exceeds it.  Choosing a richer plan costs Onyx nothing extra and costs the employee the whole difference.").font = SUB
t.cell(row=T_LAST+3, column=1, value="Medical premium only.  Admin fees, workers' compensation, setup fees, dental, vision, life and disability are excluded - see the Steven report for employer-only all-in cost.").font = SUB
for col, w in [("A", 9), ("B", 30), ("C", 36), ("D", 6), ("E", 13), ("F", 13),
               ("G", 17), ("H", 17), ("I", 17), ("J", 18), ("K", 20)]:
    t.column_dimensions[col].width = w
t.column_dimensions["C"].hidden = True
t.freeze_panes = "D5"
t.sheet_view.showGridLines = False

# =====================================================================
# 5. BY EMPLOYEE  (each enrollee against every plan)
# =====================================================================
e = wb.create_sheet("By Employee")
e["A1"] = "Every plan, priced for each Onyx enrollee"
e["A1"].font = TTL
e["A2"] = "Josephine is included for reference at family coverage even though she is waiving.  Steven's enrollment is still undecided."
e["A2"].font = SUB
E_HDR = ["Employee", "Tier", "Code", "Vendor", "Plan", "Key", "HSA",
         "Monthly premium", "Onyx $/mo", "Employee $/mo", "Employee $/yr",
         "Deductible", "OOP max", "Worst case $/yr"]
hdr(e, 4, E_HDR)
e.row_dimensions[4].height = 42
E_FIRST = 5
r = E_FIRST
for name, pay, tier, code, enrolling, _n in enr:
    for row in PLANS:
        v, plan = row[0], row[1]
        e.cell(row=r, column=1, value=name if enrolling == "Yes" else name + " (waiving)").font = INK
        e.cell(row=r, column=2, value=tier).font = INK
        e.cell(row=r, column=3, value=code).font = INK
        e.cell(row=r, column=4, value=v).font = INK
        e.cell(row=r, column=5, value=plan).font = INK
        e.cell(row=r, column=6, value=f"{v}|{plan}").font = INK
        m = f'MATCH($F{r}&"|"&$C{r},\'Cost by Tier\'!$R${CT_FIRST}:$R${CT_LAST},0)'
        for col, src in [("G", "F"), ("H", "G"), ("I", "J"), ("J", "K"),
                         ("K", "M"), ("L", "O"), ("M", "P"), ("N", "Q")]:
            e[f"{col}{r}"] = (f"=INDEX('Cost by Tier'!${src}${CT_FIRST}:${src}${CT_LAST},{m})")
            e[f"{col}{r}"].font = GRN
        for col in ("H", "I", "J"):
            e[f"{col}{r}"].number_format = CUR2
        for col in ("K", "L", "M", "N"):
            e[f"{col}{r}"].number_format = CUR
        e[f"G{r}"].alignment = Alignment(horizontal="center")
        e[f"N{r}"].font = Font(name=F, size=10, bold=True, color="008000")
        for col in range(1, 15):
            e.cell(row=r, column=col).border = BOX
        if enrolling != "Yes":
            for col in range(1, 15):
                e.cell(row=r, column=col).fill = BAND
        r += 1
E_LAST = r - 1
e.cell(row=E_LAST+2, column=1, value="Sorted by employee, then in the order the vendor lists its plans.  Use the filter on row 4 to isolate one person or one vendor.").font = SUB
for col, w in [("A", 24), ("B", 21), ("C", 6), ("D", 9), ("E", 30), ("F", 36), ("G", 6),
               ("H", 13), ("I", 12), ("J", 13), ("K", 13), ("L", 12), ("M", 12), ("N", 15)]:
    e.column_dimensions[col].width = w
e.column_dimensions["F"].hidden = True
e.freeze_panes = "G5"
e.sheet_view.showGridLines = False
e.auto_filter.ref = f"A4:N{E_LAST}"

wb.save("Onyx_2026_Medical_Plans_ADP_TriNet_Angle.xlsx")
print("saved", CT_LAST - CT_FIRST + 1, "tier rows;", E_LAST - E_FIRST + 1, "employee rows")
