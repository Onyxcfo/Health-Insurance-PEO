# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from qsheet import Q

F="Arial"
INK=Font(name=F,size=10)
BLD=Font(name=F,size=10,bold=True)
SM =Font(name=F,size=9,color="404040")
BLU=Font(name=F,size=10,color="0000FF")
TTL=Font(name=F,size=15,bold=True)
SUB=Font(name=F,size=9,italic=True,color="595959")
HDF=Font(name=F,size=9,bold=True,color="FFFFFF")
SECF=Font(name=F,size=10,bold=True,color="FFFFFF")
HFILL=PatternFill("solid",fgColor="15191C")
SECFILL=PatternFill("solid",fgColor="8E2F2A")
INPUT=PatternFill("solid",fgColor="FFFF00")
BAND=PatternFill("solid",fgColor="F2F5F6")
thin=Side(style="thin",color="BFBFBF")
BOX=Border(left=thin,right=thin,top=thin,bottom=thin)
CUR='$#,##0;($#,##0);-'
WRAP=Alignment(wrap_text=True,vertical="top")
TOP=Alignment(vertical="top")
CTR=Alignment(horizontal="center",vertical="top")

# classify the at-stake figures - summing them together would be misleading
TYPE={3240:"Hard cost in dispute",1320:"Policy choice to confirm",
      324:"Figure to verify",34800:"Contribution design basis",12000:"Conditional"}

wb=openpyxl.Workbook()

# ============================ QUESTIONS ============================
ws=wb.active; ws.title="Questions"
ws["A1"]="Questions for TriNet"; ws["A1"].font=TTL
ws["A2"]=("Onyx Accounting Group, LLC - 2026 benefits placement.  Quote Q-00412887, Aetna only, medical effective 10/1/2026.  "
          "Twenty-six items ordered by what they cost us.  Fill in the yellow columns during the meeting.")
ws["A2"].font=SUB
ws["A3"]=("Every item below is either a number we derived rather than received, a figure TriNet's own paperwork states two different ways, "
          "or a service we are being sold without a price.  Ask them in this order.")
ws["A3"].font=SUB

HDR=["#","Group","Priority","Question to ask","Why it matters","What we already hold",
     "At stake ($/yr)","Type of exposure","ANSWER","In writing?","Status","Follow-up owed"]
HR=5
for i,h in enumerate(HDR):
    c=ws.cell(row=HR,column=1+i,value=h)
    c.font=HDF; c.fill=HFILL; c.border=BOX
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
ws.row_dimensions[HR].height=34

r=HR+1
first_data=r
last_grp=None
for n,(grp,pri,q,why,held,amt,basis,fu) in enumerate(Q,start=1):
    if grp!=last_grp:                      # group banner row
        c=ws.cell(row=r,column=1,value=grp.upper())
        c.font=SECF; c.alignment=Alignment(vertical="center")
        for col in range(1,13):
            ws.cell(row=r,column=col).fill=SECFILL
            ws.cell(row=r,column=col).border=BOX
        ws.row_dimensions[r].height=20
        last_grp=grp; r+=1
    vals=[n,grp.split(". ",1)[1],pri,q,why,held,amt,(TYPE.get(amt) if amt else basis),None,None,"Open",fu]
    for j,v in enumerate(vals):
        c=ws.cell(row=r,column=1+j,value=v)
        c.border=BOX
        c.alignment=WRAP if j in (1,3,4,5,7,8,11) else TOP
        if j==0: c.font=BLD; c.alignment=CTR
        elif j==2: c.font=BLD; c.alignment=CTR
        elif j==3: c.font=BLD
        elif j in (4,5,11): c.font=SM
        elif j==6:
            c.font=BLD; c.number_format=CUR; c.alignment=Alignment(vertical="top",horizontal="right")
        elif j==7: c.font=SM; c.alignment=WRAP
        else: c.font=INK
        if j in (8,9,10): c.fill=INPUT; c.font=BLU
    ws.row_dimensions[r].height=78
    r+=1
last_data=r-1

dv_w=DataValidation(type="list",formula1='"Yes,Promised,No,N/A"',allow_blank=True,showDropDown=False)
dv_s=DataValidation(type="list",formula1='"Open,Answered,Partial,Refused,Not applicable"',allow_blank=True,showDropDown=False)
ws.add_data_validation(dv_w); ws.add_data_validation(dv_s)
dv_w.add(f"J{first_data}:J{last_data}")
dv_s.add(f"K{first_data}:K{last_data}")

ws.cell(row=last_data+2,column=1,
  value="Yellow columns are yours to fill in.  'In writing?' and 'Status' are dropdowns.  "
        "Nothing on this sheet is a formula except the tallies on the Summary tab.").font=SUB
ws.cell(row=last_data+3,column=1,
  value="Do not compare TriNet's figures for Onyx's CURRENT costs - those numbers were invented.  "
        "Keep the conversation on TriNet's own pricing.").font=Font(name=F,size=9,bold=True,italic=True,color="8E2F2A")

for col,w in [("A",5),("B",26),("C",10),("D",56),("E",50),("F",44),("G",13),("H",20),
              ("I",42),("J",11),("K",13),("L",34)]:
    ws.column_dimensions[col].width=w
ws.freeze_panes=f"D{HR+1}"
ws.sheet_view.showGridLines=False
ws.auto_filter.ref=f"A{HR}:L{last_data}"
ws.print_title_rows=f"{HR}:{HR}"

# ============================ SUMMARY ============================
s=wb.create_sheet("Summary")
s["A1"]="Where this stands"; s["A1"].font=TTL
s["A2"]="Tallies update as you fill in the Questions tab."; s["A2"].font=SUB

def sec(row,text):
    c=s.cell(row=row,column=1,value=text)
    c.font=Font(name=F,size=11,bold=True,color="8E2F2A")
    for col in range(1,5):
        s.cell(row=row,column=col).border=Border(bottom=Side(style="thin",color="8E2F2A"))

def hd(row,labels):
    for i,l in enumerate(labels):
        c=s.cell(row=row,column=1+i,value=l)
        c.font=HDF; c.fill=HFILL; c.border=BOX
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)

QCOL=f"Questions!$K${first_data}:$K${last_data}"
PCOL=f"Questions!$C${first_data}:$C${last_data}"
GCOL=f"Questions!$G${first_data}:$G${last_data}"
HCOL=f"Questions!$H${first_data}:$H${last_data}"
WCOL=f"Questions!$J${first_data}:$J${last_data}"

sec(4,"PROGRESS")
hd(5,["Status","Count","","Of the 26 questions"])
for i,st in enumerate(["Open","Answered","Partial","Refused","Not applicable"]):
    rr=6+i
    s.cell(row=rr,column=1,value=st).font=INK
    c=s.cell(row=rr,column=2,value=f'=COUNTIF({QCOL},A{rr})'); c.font=BLD
    c.alignment=Alignment(horizontal="center")
    for col in (1,2): s.cell(row=rr,column=col).border=BOX
s["D6"]='=COUNTIF('+PCOL+',"Ask first")&" are ASK FIRST - the service fee, the contribution floor and the target-date expense ratio"'
s["D6"].font=SM
s["D7"]=f'="Answered in writing: "&COUNTIF({WCOL},"Yes")'
s["D7"].font=SM

sec(13,"MONEY ON THE TABLE  (these are different kinds of exposure - do not add them together)")
hd(14,["Type of exposure","$/yr","Questions","What it means"])
TYPES=[("Hard cost in dispute","Real dollars TriNet's own documents state two different ways."),
       ("Policy choice to confirm","Ours to decide, but only once TriNet confirms $0 funding is permitted."),
       ("Figure to verify","Our re-rate of TriNet's disclosed rates. Confirm the basis."),
       ("Contribution design basis","Not at risk - but the entire $34,800 design rests on a floor rule we derived, not received."),
       ("Conditional","Only lands if Josephine enrolls at family coverage.")]
for i,(t,note) in enumerate(TYPES):
    rr=15+i
    s.cell(row=rr,column=1,value=t).font=INK
    c=s.cell(row=rr,column=2,value=f'=SUMIFS({GCOL},{HCOL},A{rr})'); c.font=BLD; c.number_format=CUR
    c=s.cell(row=rr,column=3,value=f'=COUNTIFS({HCOL},A{rr})'); c.alignment=Alignment(horizontal="center"); c.font=INK
    c=s.cell(row=rr,column=4,value=note); c.font=SM; c.alignment=WRAP
    s.row_dimensions[rr].height=28
    for col in range(1,5): s.cell(row=rr,column=col).border=BOX
rr=20
s.cell(row=rr,column=1,value="Hard cost + policy choice, combined").font=BLD
c=s.cell(row=rr,column=2,value='=B15+B16'); c.font=BLD; c.number_format=CUR
s.cell(row=rr,column=4,value="The only two lines that are genuinely undetermined dollars against a signed decision.").font=SM
for col in range(1,5): s.cell(row=rr,column=col).border=BOX

sec(23,"BEFORE YOU LEAVE THE ROOM")
CLOSERS=[("Get the floor rule and the service fee in writing, in the meeting.",
          "Everything else can follow by email. Those two decide whether TriNet is competitive at all."),
         ("Ask what happens if we leave.",
          "Notice required at a renewal, whether we get claims and loss experience data, and whether Aetna issues a certificate of prior coverage. A PEO that will not hand back experience data makes the next shopping cycle blind."),
         ("Pin the EPLI structure.",
          "$1,000,000 is stated - per claim or aggregate, and does it cover acts during the term after we terminate?"),
         ("Do not compare their figures for our current costs.",
          "Those numbers were invented. Keep the conversation on their own pricing.")]
for i,(a,b) in enumerate(CLOSERS):
    rr=24+i
    s.cell(row=rr,column=1,value=f"{i+1}.").font=BLD
    s.cell(row=rr,column=1).alignment=CTR
    c=s.cell(row=rr,column=2,value=a); c.font=BLD; c.alignment=WRAP
    c=s.cell(row=rr,column=4,value=b); c.font=SM; c.alignment=WRAP
    s.row_dimensions[rr].height=32
    for col in range(1,5): s.cell(row=rr,column=col).border=BOX
s.merge_cells(start_row=24,start_column=2,end_row=24,end_column=3)
for i in range(1,4):
    s.merge_cells(start_row=24+i,start_column=2,end_row=24+i,end_column=3)

sec(30,"SOURCES AND WHAT WE HOLD")
for i,t in enumerate([
  "TriNet quote Q-00412887 - Aetna and UHC strategies, plan designs and tier rates.",
  "TriNet VROI - service fee table and the footnote that contradicts it.",
  "TriNet 401(k) materials - Empower as recordkeeper, TriNet as plan sponsor, 0.14% average with no lineup behind it.",
  "Allison Irwin, TriNet - renewals fall on the 1st month of the quarter in which you start, so a 10/1 start holds 10/1 in perpetuity.",
  "Josephine's own notes - dental not required to be employer-funded; Electric basic tier insufficient; Perks described verbally as a 'company Groupon'.",
]):
    s.cell(row=31+i,column=1,value=t).font=SM
s.cell(row=37,column=1,
  value=("Every dollar figure was re-rated to Onyx's actual enrollment - Lisa employee-only, Christine employee+spouse, "
         "Steven and Jessica family, Josephine waiving. No vendor priced this group correctly.")).font=SUB

for col,w in [("A",34),("B",42),("C",11),("D",62)]:
    s.column_dimensions[col].width=w
s.sheet_view.showGridLines=False

wb.save("Onyx_TriNet_Questions.xlsx")
print("rows",first_data,"to",last_data,"| questions",len(Q))
